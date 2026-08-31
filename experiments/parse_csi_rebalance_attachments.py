"""Normalize official CSI rebalance PDF/XLSX attachments."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import subprocess
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

TARGETS = {"沪深300": "000300", "中证500": "000905", "中证1000": "000852"}
SECTION_RE = re.compile(r"^(沪深|中证)\s*(300|500|1000)\s*指数样本调整名单")
ROW_RE = re.compile(r"^\s*(\d{6})\s+(.+?)\s+(\d{6})\s+(.+?)\s*$")
FIELDS = (
    "notice_id",
    "publish_date",
    "effective_date",
    "index_code",
    "index_name",
    "direction",
    "security_code",
    "security_name",
    "source_file",
    "source_sha256",
)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def parse_pdf_text(text: str) -> list[dict]:
    rows = []
    index_name = None
    for raw_line in text.replace("\f", "\n").splitlines():
        line = raw_line.strip()
        section = SECTION_RE.match(line)
        if "指数样本调整名单" in line or "指数备选名单" in line:
            index_name = f"{section.group(1)}{section.group(2)}" if section else None
            continue
        match = ROW_RE.match(raw_line)
        if index_name not in TARGETS or not match:
            continue
        out_code, out_name, in_code, in_name = match.groups()
        rows.extend((
            {"index_name": index_name, "index_code": TARGETS[index_name], "direction": "OUT", "security_code": out_code, "security_name": out_name.strip()},
            {"index_name": index_name, "index_code": TARGETS[index_name], "direction": "IN", "security_code": in_code, "security_name": in_name.strip()},
        ))
    return rows


def parse_pdf(path: Path) -> list[dict]:
    result = subprocess.run(["pdftotext", "-layout", str(path), "-"], check=True, capture_output=True, text=True)
    return parse_pdf_text(result.stdout)


def parse_xlsx(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = []
    try:
        for sheet_name in ("调出", "调入"):
            if sheet_name not in workbook.sheetnames:
                continue
            direction = "OUT" if sheet_name == "调出" else "IN"
            for values in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
                if len(values) < 4 or values[0] is None or values[2] is None:
                    continue
                index_code = str(values[0]).split(".")[0].zfill(6)
                security_code = str(values[2]).split(".")[0].zfill(6)
                if index_code not in TARGETS.values():
                    continue
                index_name = next(name for name, code in TARGETS.items() if code == index_code)
                rows.append({
                    "index_name": index_name,
                    "index_code": index_code,
                    "direction": direction,
                    "security_code": security_code,
                    "security_name": str(values[3]).strip(),
                })
    finally:
        workbook.close()
    return rows


def run(capture_dir: str, out_csv: str, out_manifest: str) -> dict:
    capture = Path(capture_dir)
    output_rows = []
    unresolved = []
    for line in (capture / "notices.jsonl").read_text(encoding="utf-8").splitlines():
        notice = json.loads(line)
        effective_dates = notice["effective_dates"]
        if len(effective_dates) != 1:
            unresolved.append({"notice_id": notice["notice_id"], "reason": "effective_date_not_exact"})
            continue
        notice_rows = []
        for enclosure in notice["enclosures"]:
            path = capture / enclosure["local_file"]
            suffix = path.suffix.lower()
            parsed = parse_xlsx(path) if suffix == ".xlsx" else parse_pdf(path) if suffix == ".pdf" else []
            for row in parsed:
                notice_rows.append({
                    "notice_id": notice["notice_id"],
                    "publish_date": notice["publish_date"],
                    "effective_date": effective_dates[0],
                    **row,
                    "source_file": enclosure["local_file"],
                    "source_sha256": _sha256(path),
                })
        if not notice_rows:
            unresolved.append({"notice_id": notice["notice_id"], "reason": "no_target_rows_parsed"})
        output_rows.extend(notice_rows)
    keys = [(row["notice_id"], row["index_code"], row["direction"], row["security_code"]) for row in output_rows]
    if len(keys) != len(set(keys)):
        raise ValueError("duplicate normalized constituent change")
    out_path = Path(out_csv)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(sorted(output_rows, key=lambda row: (row["publish_date"], row["notice_id"], row["index_code"], row["direction"], row["security_code"])))
    counts = Counter((row["index_code"], row["direction"]) for row in output_rows)
    manifest = {
        "source_capture": str(capture),
        "row_count": len(output_rows),
        "counts": {f"{index_code}:{direction}": count for (index_code, direction), count in sorted(counts.items())},
        "unresolved": unresolved,
        "output_sha256": _sha256(out_path),
        "attachment_parse_complete": bool(output_rows) and not unresolved,
        "backtest_ready": False,
        "trade_authorized": False,
    }
    Path(out_manifest).write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return manifest


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--capture-dir", required=True)
    parser.add_argument("--out-csv", required=True)
    parser.add_argument("--out-manifest", required=True)
    args = parser.parse_args(argv)
    print(json.dumps(run(args.capture_dir, args.out_csv, args.out_manifest), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
